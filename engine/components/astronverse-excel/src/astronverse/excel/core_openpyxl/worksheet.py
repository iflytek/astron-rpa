from copy import copy

from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter, range_boundaries

from astronverse.excel import CopySheetLocationType, WorksheetDriverInterface
from astronverse.excel.core_openpyxl.range import OpenpyxlRangeRef
from astronverse.excel.excel_obj import ExcelObj


def _move_sheet(workbook, sheet, target_index: int):
    sheets = workbook._sheets
    sheets.remove(sheet)
    sheets.insert(target_index, sheet)


def _clone_sheet_contents(source_ws, target_ws):
    for row in source_ws.iter_rows():
        for source_cell in row:
            target_cell = target_ws.cell(row=source_cell.row, column=source_cell.column, value=source_cell.value)
            if source_cell.has_style:
                target_cell._style = copy(source_cell._style)
            if source_cell.number_format:
                target_cell.number_format = source_cell.number_format
            if source_cell.font:
                target_cell.font = copy(source_cell.font)
            if source_cell.fill:
                target_cell.fill = copy(source_cell.fill)
            if source_cell.border:
                target_cell.border = copy(source_cell.border)
            if source_cell.alignment:
                target_cell.alignment = copy(source_cell.alignment)
            if source_cell.protection:
                target_cell.protection = copy(source_cell.protection)
            if source_cell.comment:
                target_cell.comment = Comment(source_cell.comment.text, source_cell.comment.author)

    for merged_range in source_ws.merged_cells.ranges:
        target_ws.merge_cells(str(merged_range))

    for column_letter, dim in source_ws.column_dimensions.items():
        target_ws.column_dimensions[column_letter] = copy(dim)
    for row_idx, dim in source_ws.row_dimensions.items():
        target_ws.row_dimensions[row_idx] = copy(dim)


class Worksheet(WorksheetDriverInterface):
    @staticmethod
    def get_worksheet(excel_obj: ExcelObj, sheet_name: str = "", default: int = 0):
        wb = excel_obj.obj
        if not sheet_name:
            if default == 0:
                return wb.active
            sheet_name = default

        if isinstance(sheet_name, str):
            if sheet_name in wb.sheetnames:
                return wb[sheet_name]
        else:
            idx = int(sheet_name) - 1
            if 0 <= idx < len(wb.worksheets):
                return wb.worksheets[idx]
        raise ValueError(f"工作表'{sheet_name}'不存在")

    @staticmethod
    def get_all_worksheets(excel_obj: ExcelObj) -> list:
        return excel_obj.obj.worksheets

    @staticmethod
    def get_all_worksheet_names(excel_obj: ExcelObj) -> list[str]:
        return excel_obj.obj.sheetnames

    @staticmethod
    def get_active_worksheet(excel_obj: ExcelObj):
        return excel_obj.obj.active

    @staticmethod
    def add_worksheet(excel_obj: ExcelObj, sheet_name: str, before=None, after=None):
        wb = excel_obj.obj
        if before is not None:
            if isinstance(before, str):
                index = wb.sheetnames.index(before)
            else:
                index = int(before) - 1
        elif after is not None:
            if isinstance(after, str):
                index = wb.sheetnames.index(after) + 1
            else:
                index = int(after)
        else:
            index = len(wb.worksheets)
        return wb.create_sheet(title=sheet_name, index=index)

    @staticmethod
    def move_worksheet(worksheet, before=None, after=None):
        workbook = worksheet.parent
        if before is not None:
            target_index = workbook.sheetnames.index(before) if isinstance(before, str) else int(before) - 1
        elif after is not None:
            target_index = workbook.sheetnames.index(after) + 1 if isinstance(after, str) else int(after)
        else:
            target_index = len(workbook.worksheets) - 1
        _move_sheet(workbook, worksheet, max(0, min(target_index, len(workbook.worksheets) - 1)))
        workbook.active = workbook.sheetnames.index(worksheet.title)

    @staticmethod
    def get_worksheet_name(worksheet) -> str:
        return worksheet.title

    @staticmethod
    def rename_worksheet(worksheet, new_name: str):
        worksheet.title = new_name

    @staticmethod
    def delete_worksheet(worksheet):
        worksheet.parent.remove(worksheet)

    @staticmethod
    def copy_worksheet(
        worksheet, excel, location: CopySheetLocationType = CopySheetLocationType.LAST, is_same_workbook=False
    ):
        source_wb = worksheet.parent
        target_wb = excel.obj

        if is_same_workbook:
            new_sheet = source_wb.copy_worksheet(worksheet)
        else:
            new_sheet = target_wb.create_sheet(title=f"{worksheet.title}_copy")
            _clone_sheet_contents(worksheet, new_sheet)

        active_title = target_wb.active.title if target_wb.worksheets else new_sheet.title
        if location == CopySheetLocationType.BEFORE:
            target_index = (
                target_wb.sheetnames.index(worksheet.title)
                if is_same_workbook
                else target_wb.sheetnames.index(active_title)
            )
        elif location == CopySheetLocationType.AFTER:
            target_index = (
                target_wb.sheetnames.index(worksheet.title) + 1
                if is_same_workbook
                else target_wb.sheetnames.index(active_title) + 1
            )
        elif location == CopySheetLocationType.FIRST:
            target_index = 0
        else:
            target_index = len(target_wb.worksheets) - 1

        _move_sheet(target_wb, new_sheet, target_index)
        target_wb.active = target_wb.sheetnames.index(new_sheet.title)
        return new_sheet

    @staticmethod
    def get_worksheet_used_range(worksheet):
        if worksheet.max_row == 1 and worksheet.max_column == 1 and worksheet["A1"].value is None:
            return 1, 1, 0, 0, "A1"
        min_col, min_row, max_col, max_row = range_boundaries(worksheet.calculate_dimension())
        address = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
        return min_row, min_col, max_row, max_col, address

    @staticmethod
    def get_cell(worksheet, row: int, col: int):
        coordinate = f"{get_column_letter(col)}{row}"
        return OpenpyxlRangeRef.from_address(worksheet, coordinate)

    @staticmethod
    def get_range(worksheet, cell: str):
        return OpenpyxlRangeRef.from_address(worksheet, cell)

    @staticmethod
    def get_rows(worksheet, rows):
        return OpenpyxlRangeRef.for_rows(worksheet, rows)

    @staticmethod
    def get_columns(worksheet, columns):
        return OpenpyxlRangeRef.for_columns(worksheet, columns)

    @staticmethod
    def get_range_from_cells(worksheet, start_cell, end_cell):
        start = start_cell.address.split(":")[0]
        end = end_cell.address.split(":")[0]
        return OpenpyxlRangeRef.from_address(worksheet, f"{start}:{end}")

    @staticmethod
    def insert_picture(worksheet, image_path, pic_left=0, pic_top=0, pic_height=300, pic_width=400, pic_scale=1.0):
        raise NotImplementedError("文件路径有误，请输入正确的路径！")

    @staticmethod
    def delete_all_comments(worksheet):
        found = False
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.comment:
                    cell.comment = None
                    found = True
        if not found:
            raise ValueError("不存在批注")

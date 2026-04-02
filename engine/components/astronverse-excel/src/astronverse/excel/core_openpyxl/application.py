import os
from typing import Optional

from openpyxl import Workbook, load_workbook

from astronverse.excel import ApplicationDriverInterface, ApplicationType
from astronverse.excel.error import BizException, DRIVER_UNSUPPORTED_ERROR, FILE_PATH_ERROR
from astronverse.excel.excel_obj import ExcelObj


class Application(ApplicationDriverInterface):
    APP_MARKER = "__OPENPYXL_APP__"
    SUPPORTED_SUFFIXES = {".xlsx", ".xlsm", ".xltx", ".xltm"}

    @staticmethod
    def _ensure_supported_path(file_path: str):
        suffix = os.path.splitext(file_path)[1].lower()
        if suffix not in Application.SUPPORTED_SUFFIXES:
            raise BizException(FILE_PATH_ERROR, f"文件路径有误，请输入正确的路径！: {file_path}")

    @staticmethod
    def _ensure_password_supported(password: str = "", write_res_password: str = ""):
        if password or write_res_password:
            raise BizException(DRIVER_UNSUPPORTED_ERROR, "当前驱动程序不支持此功能")

    @staticmethod
    def init_app(
        default_application: ApplicationType = ApplicationType.OPENPYXL,
        visible_flag: bool = None,
        retry: int = 0,
        retry_delay: float = 0.5,
        prefer_existing: bool = True,
    ) -> str:
        return Application.APP_MARKER

    @staticmethod
    def quit_app(default_application: ApplicationType = ApplicationType.DEFAULT, save_changes: bool = False):
        return None

    @staticmethod
    def open_workbook(
        application: Optional[object],
        file_path: str,
        password: str = "",
        update_links: bool = True,
        write_res_password: str = "",
    ) -> ExcelObj:
        Application._ensure_password_supported(password=password, write_res_password=write_res_password)
        Application._ensure_supported_path(file_path)
        wb = load_workbook(filename=file_path, data_only=False, keep_vba=file_path.lower().endswith(".xlsm"))
        return ExcelObj(obj=wb, path=file_path, app_type=ApplicationType.OPENPYXL)

    @staticmethod
    def create_workbook(application: Optional[object], file_path: str = "", password: str = "") -> ExcelObj:
        Application._ensure_password_supported(password=password)
        wb = Workbook()
        excel_obj = ExcelObj(obj=wb, path=file_path, app_type=ApplicationType.OPENPYXL)
        if file_path:
            Application._ensure_supported_path(file_path)
            wb.save(file_path)
        return excel_obj

    @staticmethod
    def get_existing_workbook(application, match_name: str):
        raise BizException(DRIVER_UNSUPPORTED_ERROR, "当前驱动程序不支持此功能")

    @staticmethod
    def save_workbook(excel_obj: ExcelObj, file_path: str = "", password: str = ""):
        Application._ensure_password_supported(password=password)
        save_path = file_path or excel_obj.path
        if not save_path:
            raise BizException(FILE_PATH_ERROR, "文件路径有误，请输入正确的路径！")
        Application._ensure_supported_path(save_path)
        excel_obj.obj.save(save_path)
        excel_obj.path = save_path

    @staticmethod
    def close_workbook(excel_obj: ExcelObj, save_changes: bool = False):
        if save_changes and excel_obj.path:
            excel_obj.obj.save(excel_obj.path)
        excel_obj.obj.close()

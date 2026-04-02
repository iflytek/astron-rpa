from copy import copy
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from typing import Any

from openpyxl.comments import Comment
from openpyxl.formula.translate import Translator
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter, range_boundaries

from astronverse.excel import (
    RangeDriverInterface,
    ClearType,
    FontNameType,
    FontType,
    HorizontalAlign,
    MergeOrSplitType,
    NumberFormatType,
    ReadRangeType,
    SetType,
    VerticalAlign,
)


@dataclass
class OpenpyxlRangeRef:
    worksheet: Any
    min_row: int
    min_col: int
    max_row: int
    max_col: int
    address: str
    kind: str = "range"

    @classmethod
    def from_address(cls, worksheet, address: str):
        min_col, min_row, max_col, max_row = range_boundaries(address)
        if min_row == max_row and min_col == max_col:
            kind = "cell"
        elif min_col == 1 and max_col == (worksheet.max_column or max_col):
            kind = "row"
        elif min_row == 1 and max_row == (worksheet.max_row or max_row):
            kind = "column"
        else:
            kind = "range"
        return cls(worksheet, min_row, min_col, max_row, max_col, address, kind=kind)

    @classmethod
    def for_rows(cls, worksheet, rows):
        return cls(
            worksheet,
            int(rows),
            1,
            int(rows),
            worksheet.max_column or 1,
            f"A{rows}:{get_column_letter(worksheet.max_column or 1)}{rows}",
            kind="row",
        )

    @classmethod
    def for_columns(cls, worksheet, columns):
        col_num = int(columns) if str(columns).isdigit() else column_index_from_string(str(columns))
        return cls(
            worksheet,
            1,
            col_num,
            worksheet.max_row or 1,
            col_num,
            f"{get_column_letter(col_num)}1:{get_column_letter(col_num)}{worksheet.max_row or 1}",
            kind="column",
        )

    def iter_cells(self):
        for row in self.worksheet.iter_rows(
            min_row=self.min_row, max_row=self.max_row, min_col=self.min_col, max_col=self.max_col
        ):
            for cell in row:
                yield cell

    def iter_row_groups(self):
        for row in self.worksheet.iter_rows(
            min_row=self.min_row, max_row=self.max_row, min_col=self.min_col, max_col=self.max_col
        ):
            yield row

    def first_cell(self):
        return self.worksheet.cell(self.min_row, self.min_col)

    @property
    def row_count(self) -> int:
        return self.max_row - self.min_row + 1

    @property
    def col_count(self) -> int:
        return self.max_col - self.min_col + 1


_OPENPYXL_CLIPBOARD = None


def _to_text(value):
    if value is None:
        return None
    return str(value)


def _parse_rgb(color_value: str):
    color_value = (color_value or "").replace("FF", "", 1)
    if len(color_value) != 6:
        return 255, 255, 255
    return int(color_value[0:2], 16), int(color_value[2:4], 16), int(color_value[4:6], 16)


def _clone_cell(source_cell, target_cell, copy_value=True, copy_style=True, copy_comment=True, skip_blanks=False):
    if copy_value and not (skip_blanks and source_cell.value in ("", None)):
        target_cell.value = source_cell.value
    if copy_style:
        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)
        target_cell.font = copy(source_cell.font)
        target_cell.fill = copy(source_cell.fill)
        target_cell.border = copy(source_cell.border)
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.number_format = source_cell.number_format
        target_cell.protection = copy(source_cell.protection)
    if copy_comment:
        target_cell.comment = (
            Comment(source_cell.comment.text, source_cell.comment.author) if source_cell.comment else None
        )


class Range(RangeDriverInterface):
    @staticmethod
    def get_range_data(range_obj: OpenpyxlRangeRef, use_text: bool = False) -> Any:
        if range_obj.kind == "cell":
            value = range_obj.first_cell().value
            return _to_text(value) if use_text else value
        data = []
        for row in range_obj.iter_row_groups():
            row_data = []
            for cell in row:
                row_data.append(_to_text(cell.value) if use_text else cell.value)
            data.append(row_data)
        return data

    @staticmethod
    def get_range_color(range_obj: OpenpyxlRangeRef) -> tuple[int, int, int]:
        cell = range_obj.first_cell()
        rgb = getattr(cell.fill.fgColor, "rgb", None)
        return _parse_rgb(rgb)

    @staticmethod
    def get_range_size(range_obj: OpenpyxlRangeRef) -> tuple[int, int, int, int]:
        raise NotImplementedError("Openpyxl does not support COM-style pixel geometry")

    @staticmethod
    def set_range_data(range_obj: OpenpyxlRangeRef, value: Any):
        if range_obj.kind == "cell":
            range_obj.first_cell().value = value
            return

        if not isinstance(value, list):
            for cell in range_obj.iter_cells():
                cell.value = value
            return

        for r_offset, row in enumerate(value):
            if not isinstance(row, list):
                row = [row]
            for c_offset, cell_value in enumerate(row):
                target_row = range_obj.min_row + r_offset
                target_col = range_obj.min_col + c_offset
                if target_row <= range_obj.max_row and target_col <= range_obj.max_col:
                    range_obj.worksheet.cell(target_row, target_col, cell_value)

    @staticmethod
    def set_range_type(
        range_obj: OpenpyxlRangeRef,
        col_width=None,
        bg_color=None,
        font_color=None,
        font_type: FontType = FontType.NO_CHANGE,
        font_name: FontNameType = FontNameType.NO_CHANGE,
        font_size=None,
        number_format: NumberFormatType = NumberFormatType.NO_CHANGE,
        number_format_other: str = "",
        horizontal_align: HorizontalAlign = HorizontalAlign.NO_CHANGE,
        vertical_align: VerticalAlign = VerticalAlign.NO_CHANGE,
        wrap_text: bool = True,
        design_type: ReadRangeType = ReadRangeType.CELL,
        auto_row_height: bool = False,
        auto_column_width: bool = False,
    ):
        h_align_map = {
            HorizontalAlign.DEFAULT: "general",
            HorizontalAlign.LEFT: "left",
            HorizontalAlign.RIGHT: "right",
            HorizontalAlign.CENTER: "center",
            HorizontalAlign.PADDING: "fill",
            HorizontalAlign.BOTH: "justify",
            HorizontalAlign.CROSS: "centerContinuous",
            HorizontalAlign.DISTRIBUTED: "distributed",
        }
        v_align_map = {
            VerticalAlign.UP: "top",
            VerticalAlign.MIDDLE: "center",
            VerticalAlign.DOWN: "bottom",
            VerticalAlign.BOTH: "justify",
            VerticalAlign.DISTRIBUTED: "distributed",
        }

        for cell in range_obj.iter_cells():
            font = copy(cell.font) if cell.font else Font()
            if font_name != FontNameType.NO_CHANGE:
                font.name = font_name.value
            if font_size:
                font.size = font_size
            if font_type == FontType.BOLD:
                font.bold = True
            elif font_type == FontType.ITALIC:
                font.italic = True
            elif font_type == FontType.BOLD_ITALIC:
                font.bold = True
                font.italic = True
            elif font_type == FontType.NORMAL:
                font.bold = False
                font.italic = False
            if font_color:
                font.color = "".join(f"{c:02X}" for c in font_color)
            cell.font = font

            if bg_color:
                rgb = "".join(f"{c:02X}" for c in bg_color)
                cell.fill = PatternFill(fill_type="solid", fgColor=rgb)

            alignment = copy(cell.alignment)
            alignment = Alignment(
                horizontal=h_align_map.get(horizontal_align, alignment.horizontal)
                if horizontal_align != HorizontalAlign.NO_CHANGE
                else alignment.horizontal,
                vertical=v_align_map.get(vertical_align, alignment.vertical)
                if vertical_align != VerticalAlign.NO_CHANGE
                else alignment.vertical,
                wrap_text=wrap_text,
            )
            cell.alignment = alignment

            if number_format != NumberFormatType.NO_CHANGE:
                cell.number_format = (
                    number_format_other if number_format == NumberFormatType.CUSTOM else number_format.value
                )

        if col_width:
            for col_idx in range(range_obj.min_col, range_obj.max_col + 1):
                range_obj.worksheet.column_dimensions[get_column_letter(col_idx)].width = float(col_width)

        if design_type == ReadRangeType.ROW and auto_row_height:
            for row_idx in range(range_obj.min_row, range_obj.max_row + 1):
                range_obj.worksheet.row_dimensions[row_idx].height = None
        if design_type == ReadRangeType.COLUMN and auto_column_width:
            for col_idx in range(range_obj.min_col, range_obj.max_col + 1):
                letter = get_column_letter(col_idx)
                max_length = 0
                for row_idx in range(1, range_obj.worksheet.max_row + 1):
                    value = range_obj.worksheet.cell(row_idx, col_idx).value
                    max_length = max(max_length, len(str(value)) if value is not None else 0)
                range_obj.worksheet.column_dimensions[letter].width = max(max_length + 2, 8)

    @staticmethod
    def delete_range(range_obj: OpenpyxlRangeRef, direction: str = "") -> None:
        ws = range_obj.worksheet
        if range_obj.kind == "row":
            ws.delete_rows(range_obj.min_row, range_obj.row_count)
            return
        if range_obj.kind == "column":
            ws.delete_cols(range_obj.min_col, range_obj.col_count)
            return

        if direction == "lower_move_up":
            shift = range_obj.row_count
            for row in range(range_obj.min_row, ws.max_row - shift + 1):
                for col in range(range_obj.min_col, range_obj.max_col + 1):
                    ws.cell(row, col).value = ws.cell(row + shift, col).value
                    ws.cell(row, col)._style = copy(ws.cell(row + shift, col)._style)
            for row in range(ws.max_row - shift + 1, ws.max_row + 1):
                for col in range(range_obj.min_col, range_obj.max_col + 1):
                    ws.cell(row, col).value = None
        elif direction == "right_move_left":
            shift = range_obj.col_count
            for col in range(range_obj.min_col, ws.max_column - shift + 1):
                for row in range(range_obj.min_row, range_obj.max_row + 1):
                    ws.cell(row, col).value = ws.cell(row, col + shift).value
                    ws.cell(row, col)._style = copy(ws.cell(row, col + shift)._style)
            for col in range(ws.max_column - shift + 1, ws.max_column + 1):
                for row in range(range_obj.min_row, range_obj.max_row + 1):
                    ws.cell(row, col).value = None
        else:
            for cell in range_obj.iter_cells():
                cell.value = None

    @staticmethod
    def clear_range(range_obj: OpenpyxlRangeRef, clear_type: str = ""):
        for cell in range_obj.iter_cells():
            if clear_type == ClearType.CONTENT.value:
                cell.value = None
            elif clear_type == ClearType.STYLE.value:
                cell.font = Font()
                cell.fill = PatternFill()
                cell.border = copy(cell.border.__class__())
                cell.alignment = Alignment()
                cell.number_format = "General"
                cell.comment = None
            elif clear_type == ClearType.ALL.value:
                cell.value = None
                cell.font = Font()
                cell.fill = PatternFill()
                cell.border = copy(cell.border.__class__())
                cell.alignment = Alignment()
                cell.number_format = "General"
                cell.comment = None
            else:
                raise ValueError(f"不支持的清理类型: {clear_type}")

    @staticmethod
    def copy_range(range_obj: OpenpyxlRangeRef):
        global _OPENPYXL_CLIPBOARD
        cells = []
        for row in range_obj.iter_row_groups():
            cell_row = []
            for cell in row:
                cell_row.append(
                    {
                        "value": cell.value,
                        "style": copy(cell._style),
                        "font": copy(cell.font),
                        "fill": copy(cell.fill),
                        "border": copy(cell.border),
                        "alignment": copy(cell.alignment),
                        "number_format": cell.number_format,
                        "comment": Comment(cell.comment.text, cell.comment.author) if cell.comment else None,
                    }
                )
            cells.append(cell_row)
        _OPENPYXL_CLIPBOARD = {"cells": cells, "rows": range_obj.row_count, "cols": range_obj.col_count}

    @staticmethod
    def paste_range(range_obj: OpenpyxlRangeRef, paste_type: str = "", skip_blanks=False, transpose=False):
        global _OPENPYXL_CLIPBOARD
        if not _OPENPYXL_CLIPBOARD:
            return
        paste_type = paste_type or "all"

        cells = _OPENPYXL_CLIPBOARD["cells"]
        if transpose:
            cells = [list(row) for row in zip(*cells)]

        value_only = paste_type in ("paste_value",)
        format_only = paste_type in ("format", "col_width_only")
        copy_style = paste_type in (
            "all",
            "value_and_format",
            "format",
            "formula_and_format",
            "exclude_frame",
            "col_width_only",
        )
        copy_value = paste_type in ("all", "value_and_format", "paste_value", "formula_only", "formula_and_format")

        for r_offset, row in enumerate(cells):
            for c_offset, source in enumerate(row):
                target = range_obj.worksheet.cell(range_obj.min_row + r_offset, range_obj.min_col + c_offset)
                if format_only and paste_type == "col_width_only":
                    letter = get_column_letter(range_obj.min_col + c_offset)
                    range_obj.worksheet.column_dimensions[letter].width = range_obj.worksheet.column_dimensions[
                        letter
                    ].width
                    continue
                if skip_blanks and source["value"] in ("", None):
                    continue
                if copy_value and not value_only:
                    target.value = source["value"]
                elif value_only:
                    target.value = source["value"]
                if copy_style:
                    target._style = copy(source["style"])
                    target.font = copy(source["font"])
                    target.fill = copy(source["fill"])
                    target.border = copy(source["border"])
                    target.alignment = copy(source["alignment"])
                    target.number_format = source["number_format"]
                target.comment = (
                    Comment(source["comment"].text, source["comment"].author) if source["comment"] else None
                )

    @staticmethod
    def insert_range(range_obj: OpenpyxlRangeRef, axis: str = "row"):
        if axis == "row":
            range_obj.worksheet.insert_rows(range_obj.min_row, amount=range_obj.row_count)
        elif axis == "column":
            range_obj.worksheet.insert_cols(range_obj.min_col, amount=range_obj.col_count)
        else:
            raise ValueError(f"不支持的axis参数: {axis}")

    @staticmethod
    def merge_range(range_obj: OpenpyxlRangeRef, job_type: str):
        if job_type == MergeOrSplitType.MERGE.value:
            range_obj.worksheet.merge_cells(range_obj.address)
        else:
            range_obj.worksheet.unmerge_cells(range_obj.address)

    @staticmethod
    def autofill_range(range_obj: OpenpyxlRangeRef, target_range: OpenpyxlRangeRef):
        source = range_obj.first_cell()
        for row in range(target_range.min_row, target_range.max_row + 1):
            for col in range(target_range.min_col, target_range.max_col + 1):
                target = target_range.worksheet.cell(row, col)
                if row == source.row and col == source.column:
                    continue
                if isinstance(source.value, str) and source.value.startswith("="):
                    target.value = Translator(source.value, origin=source.coordinate).translate_formula(
                        target.coordinate
                    )
                else:
                    target.value = source.value
                target._style = copy(source._style)

    @staticmethod
    def set_row_height(range_obj: OpenpyxlRangeRef, set_type: SetType, height_float: float):
        for row_idx in range(range_obj.min_row, range_obj.max_row + 1):
            range_obj.worksheet.row_dimensions[row_idx].height = height_float if set_type == SetType.VALUE else None

    @staticmethod
    def set_column_width(range_obj: OpenpyxlRangeRef, set_type: SetType, width_float: float):
        for col_idx in range(range_obj.min_col, range_obj.max_col + 1):
            letter = get_column_letter(col_idx)
            if set_type == SetType.VALUE:
                range_obj.worksheet.column_dimensions[letter].width = width_float
            else:
                max_length = 0
                for row_idx in range(1, range_obj.worksheet.max_row + 1):
                    value = range_obj.worksheet.cell(row_idx, col_idx).value
                    max_length = max(max_length, len(str(value)) if value is not None else 0)
                range_obj.worksheet.column_dimensions[letter].width = max(max_length + 2, 8)

    @staticmethod
    def convert_text_to_number(range_obj: OpenpyxlRangeRef, temp_range):
        for cell in range_obj.iter_cells():
            if isinstance(cell.value, str):
                raw = cell.value.strip().replace(",", "")
                try:
                    number = Decimal(raw)
                except InvalidOperation:
                    continue
                cell.value = int(number) if number == int(number) else float(number)

    @staticmethod
    def convert_number_to_text(range_obj: OpenpyxlRangeRef):
        for cell in range_obj.iter_cells():
            if cell.value is not None and not isinstance(cell.value, str):
                cell.number_format = "@"
                cell.value = str(cell.value)

    @staticmethod
    def add_comment(range_obj: OpenpyxlRangeRef, comment_text: str):
        range_obj.first_cell().comment = Comment(comment_text, "astronverse")

    @staticmethod
    def delete_comment(range_obj: OpenpyxlRangeRef):
        cell = range_obj.first_cell()
        if not cell.comment:
            raise ValueError("不存在批注")
        cell.comment = None

    @staticmethod
    def search_and_replace(
        range_obj: OpenpyxlRangeRef,
        find_str: str,
        replace_str: str = "",
        exact_match: bool = False,
        case_flag: bool = False,
        match_all: bool = True,
    ) -> list:
        results = []
        needle = find_str if case_flag else find_str.lower()
        for cell in range_obj.iter_cells():
            value = "" if cell.value is None else str(cell.value)
            compare_value = value if case_flag else value.lower()
            matched = compare_value == needle if exact_match else needle in compare_value
            if not matched:
                continue

            results.append({"row": str(cell.row), "col": get_column_letter(cell.column)})
            if replace_str:
                cell.value = value.replace(find_str, replace_str)
            if not match_all:
                break
        return results
